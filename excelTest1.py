import openpyxl
from openpyxl import Workbook

# 엑셀파일 열기
wb1 = openpyxl.load_workbook('2020년 보안 QUIZ 참여자 리스트_v0.1.xlsx')
wb2 = openpyxl.load_workbook('20200715_정답회신리스트.xlsx')
wb3 = Workbook()
file_name = '합계.xlsx'
ws3 = wb3.active
ws3.title = '합계'

# 현재 Active Sheet 얻기
ws1 = wb1.active

# ws = wb.get_sheet_by_name("Sheet1")

a = []
b = []
for r1 in ws1.rows:
    if r1[0].row > 2:
        name = r1[4].value
        a.append(name)
        if r1[0].row > 200:
            break

ws2 = wb2.active

for r1 in ws2.rows:
    if r1[0].row > 1:
        name = r1[6].value
        b.append(name)
        if r1[0].row > 200:
            break

c = a+b

new_list = []
for v in c:
    if v not in new_list:
        new_list.append(v)



for item in new_list:
    print(item)
    total = 0

    for r1 in ws1.rows:
        if r1[0].row > 2:
            if r1[0].row > 200:
                break
            w1personName = r1[4].value
            w1score = 0
            if 'int' == type(r1[5].value) and r1[5].value != 'None':
                w1score = r1[5].value
            elif r1[5].value is not None:
                w1score = int(r1[5].value)
            if item == w1personName:
                total = total + w1score


    for r2 in ws2.rows:
        if r2[0].row > 1:
            if r2[0].row > 200:
                break
            w2personName = r2[6].value
            w2score = 0
            if 'int' == type(r2[8].value):
                w2score = r2[8].value
            elif r2[8].value is not None:
                w2score = int(r2[8].value)
            if item == w2personName:
                # print(total)
                # print(type(r2[8].value))
                # print(w2score)
                total = total + w2score

    row_index = new_list.index(item) + 3
    ws3.cell(row=row_index, column=5).value = item
    ws3.cell(row=row_index, column=6).value = total

    # row_index = r1[0].row
    # w1personName = r1[4].value
    # w1score = r1[5].value
    # for r2 in ws2.rows:
    #     w2personName = r2[4].value
    #     w2score = r2[5].value
    #     if w1personName > w2personName:
    #         ws1.cell(row=row_index, column=6).value = w1personName
    #         ws1.cell(row=row_index, column=7).value = w1score + w2score
    # row_index = r[0].row  # 행 인덱스
    # kor = r[1].value
    # eng = r[2].value
    # math = r[3].value
    # sum = kor + eng + math
    # # 합계 쓰기
    # ws1.cell(row=row_index, column=5).value = sum
wb1.close()
wb2.close()
# 엑셀 파일 저장
wb3.save(filename=file_name)
wb3.close()